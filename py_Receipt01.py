#-*- coding:utf-8 -*-

import os
from pdf2image import convert_from_path

import re
from io import StringIO
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

import openpyxl
import datetime
from dateutil.parser import parse # for 날짜시간 format
#from operator import itemgetter # for sort list of lists?? 없어도 되는 듯??
from copy import copy # cell.style copy할 때 필요함

# name of jpg
output_string = StringIO()
strings_byPage = []
with open('input.pdf', 'rb') as in_file:
    parser = PDFParser(in_file)
    doc = PDFDocument(parser)
    rsrcmgr = PDFResourceManager()
    device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    for page in PDFPage.create_pages(doc):
        interpreter.process_page(page)

        strings_byPage.append(re.sub('\n\n', '\n', output_string.getvalue().strip()))

        output_string.seek(0)
        output_string.truncate(0)

lines_xlsm = []

# pdf to jpg
file_name = "input.pdf"
pages = convert_from_path(file_name)
os.makedirs('./JPGs', exist_ok=True)
for i, page in enumerate(pages):
    strings_byLine = strings_byPage[i].split('\n')
    dateTime = strings_byLine[3].replace('/', '.')
    dateTime = dateTime.replace(':', '.')
    page.save("./JPGs/" + dateTime +".jpg", "JPEG")
    
    # xlsm 작업 준비
    line_xlsm = []
    line_xlsm.append('비씨6666')
    date = parse(strings_byLine[3])
    line_xlsm.append(date)
    line_xlsm.append('최강그룹')
    line_xlsm.append('홍길동')
    line_xlsm.append('복리후생비')  #??
    line_xlsm.append(strings_byLine[10])
    line_xlsm.append(1)  #??
    line_xlsm.append('########')  #??
    line_xlsm.append(strings_byLine[6])
    lines_xlsm.append(line_xlsm)

# sort(ordering)
# sort(스스로), sorted(반환값)
lines_xlsm.sort(key = lambda x:x[1])

# xlsm
def insert_rows(worksheet, num):
    for i in range(num):
        row_taget = 27 + i
        row_taget += 1
        worksheet.insert_rows(row_taget)
        for row in worksheet['A27:Z27']:
            for cell in row:
                cell_target = worksheet.cell(row=row_taget, column=cell.column)
                if type(cell_target).__name__ == 'MergedCell':
                    continue
                cell_target.data_type = cell.data_type
                if cell.has_style:
                    cell_target.style = copy(cell.style)
                    cell_target.font = copy(cell.font)
                    cell_target.border = copy(cell.border)
                    cell_target.fill = copy(cell.fill)
                    cell_target.number_format = copy(cell.number_format)
                    cell_target.protection = copy(cell.protection)
                    cell_target.alignment = copy(cell.alignment)

        # openpyxl 넘버링 = 1부터 시작
        for merged in ws.merged_cells.ranges:
            if len(merged.left) == 1 and merged.start_cell.row == 27:
                coord = merged.coord.replace('27', str(row_taget))
                worksheet.merge_cells(coord)

filename_input = "영수증처리사용내역서_yyyymmdd.xlsm"
wb = openpyxl.load_workbook(filename_input, keep_vba=True)  # Workbook 객체 생성
ws = wb['법인카드']

now = datetime.datetime.now()
date = f'{now.year}{now.month:02}{now.day:02}'
filename_output = filename_input.replace('yyyymmdd', date)

for row in ws['A7:Z27']:  # 범위 지우기(MergedCell)
    for cell in row:
        if type(cell).__name__ != 'MergedCell':
            cell.value = None

insert_rows(ws, len(lines_xlsm))

for i, line_xlsm in enumerate(lines_xlsm):
    row = i + 7
    #print(','.join([str(f'{i:02}'), str(line_xlsm)]))

    line_xlsm.append(line_xlsm[1])  # 내가 참고하려고 비고에 추가함
    dateTime = line_xlsm[1]
    line_xlsm[1] = datetime.datetime(dateTime.year, dateTime.month, dateTime.day)
    
    cnt_line_xlsm = 0
    coord = 'A7:Z7'.replace('7', str(row))
    for row in ws[coord]:
        for cell in row:
            if type(cell).__name__ != 'MergedCell' and cnt_line_xlsm < len(line_xlsm):
                cell.value = line_xlsm[cnt_line_xlsm]
                cnt_line_xlsm += 1

wb.save(filename_output)
